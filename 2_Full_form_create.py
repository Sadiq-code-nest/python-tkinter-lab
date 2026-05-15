# Full Tkinter Form Application with Excel Storage

import tkinter
from tkinter import ttk
from tkinter import messagebox
import openpyxl
import os

root = tkinter.Tk()
root.title("Student Registration Form")
root.geometry("400x250")
root.resizable(False, False)

file_path = "/home/cloudly/python-tkinter-lab/Tkinter-test.xlsx"

# Create workbook if file does not exist
if not os.path.exists(file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Create Headers
    sheet["A1"] = "Name"
    sheet["B1"] = "Email"
    sheet["C1"] = "Department"

    workbook.save(file_path)

# Load workbook
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Submit Function

def submit_data():

    name = name_entry.get()
    email = email_entry.get()
    department = department_dropdown.get()

    # Validation
    if name == "" or email == "" or department == "":
        messagebox.showerror(
            "Error",
            "All fields are required!"
        )
        return

    # Find next row
    next_row = sheet.max_row + 1

    # Save Data
    sheet.cell(row=next_row, column=1).value = name
    sheet.cell(row=next_row, column=2).value = email
    sheet.cell(row=next_row, column=3).value = department

    # Save Workbook
    workbook.save(file_path)

    # Success Message
    messagebox.showinfo(
        "Success",
        "Data submitted successfully!"
    )

    # Clear Fields
    name_entry.delete(0, tkinter.END)
    email_entry.delete(0, tkinter.END)
    department_dropdown.set("Select Department")

# Label

title_label = tkinter.Label(
    root,
    text="Student Registration Form",
    font=("Arial", 14, "bold")
)
title_label.grid(row=0, column=0, columnspan=2, pady=10)

# Name
name_label = tkinter.Label(root, text="Full Name:")
name_label.grid(row=1, column=0, padx=10, pady=5, sticky="w")

name_entry = tkinter.Entry(root, width=30)
name_entry.grid(row=1, column=1, padx=10, pady=5)

# Email
email_label = tkinter.Label(root, text="Email:")
email_label.grid(row=2, column=0, padx=10, pady=5, sticky="w")

email_entry = tkinter.Entry(root, width=30)
email_entry.grid(row=2, column=1, padx=10, pady=5)

# Department Dropdown
department_label = tkinter.Label(root, text="Department:")
department_label.grid(row=3, column=0, padx=10, pady=5, sticky="w")

department_options = [
    "CSE",
    "EEE",
    "ECE",
    "IPE",
    "MECH",
    "CIVIL",
    "CHEM",
    "BIOTECH",
    "AERO"
]

department_dropdown = ttk.Combobox(
    root,
    values=department_options,
    width=27
)

department_dropdown.grid(row=3, column=1, padx=10, pady=5)
department_dropdown.set("Select Department")

# Submit Button
submit_button = tkinter.Button(
    root,
    text="Submit",
    command=submit_data,
    width=20,
    bg="green",
    fg="white"
)

submit_button.grid(row=4, column=0, columnspan=2, pady=20)

root.mainloop()